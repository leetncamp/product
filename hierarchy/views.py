from django.shortcuts import render, redirect
from django.core.urlresolvers import reverse
from django import forms
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from hierarchy.models import *
import traceback
from django.db import IntegrityError
from django import forms
from tempfile import NamedTemporaryFile
from django.http import HttpResponse, JsonResponse
import os, sys
from django.contrib import messages
from django.conf import settings
import base64
import subprocess
import time
import shutil
import glob

class SubProductForm(forms.Form):
    subproduct = forms.FileField()

class ProductForm(forms.Form):
    product = forms.FileField()

def new(request):
    newcodes = []
    subproductform = SubProductForm()
    download = "false"
    productform = ProductForm()
    if request.method == "POST":
        if request.POST.get("select", "") == "subproduct":
            wb = load_workbook(request.FILES['subproduct'])
            ws = wb.active
            headers = [cell.value for cell in ws.rows.next() ]
            subprods=[]
            for row in ws.iter_rows(row_offset=1):
                if row[0].value:
                    rowDict = {}
                    for cellnum in range(len(row) + 1):
                        try:
                            cell = row[cellnum]
                            rowDict[headers[cellnum]] = cell.value
                        except IndexError:
                            pass
                    try:
                        productline = ProductLine.objects.get(code=rowDict.get("Existing Product Line Code"))
                    except Exception as e:
                        print(u"Could not find ProductLine '{0}' at row {1}".format(rowDict.get("Existing Product Line Code"), row[0].row))
                        debug()

                    code = get_unused_code()
                    description = rowDict.get("Igor / Sub PL Description", "")
                    usage = rowDict.get("Usage")
                    if usage:
                        usage = Usage.objects.get(name=usage)
                    igoritemclass = rowDict.get("Igor Item Class", None)
                    if igoritemclass:
                        igoritemclass = IgorItemClass.objects.get(name=igoritemclass)
                    if len(description) >  30:
                        print(u"New Subproduct Line '{0}' exceed 30 characters: {0}. Dropping into the debugger. You decide what to do.".format(description, len(description)))
                        debug()

                    subproductline, created = SubProductLine.objects.get_or_create(fproductline=productline,
                        igor_or_sub_pl=code.code, description=description, igorclass=igoritemclass, usage=usage)
                    if created:
                        msg = subproductline.save()
                        if msg:
                            messages.warning(request, msg)
                        subprods.append(subproductline)
                        code.use(newcodes)


            wb = Workbook()
            ws = wb.active
            count=1
            tmp = NamedTemporaryFile(suffix=".xlsx")
            for header in bigheaders:
                ws.cell(row=1, column=count).value= header
                count += 1
            count = 2

            for subprod in subprods:
                subprod.excel_row(ws, count)
                count += 1
            wb.save(tmp)
            tmp.flush()
            #tmp.seek(0)
            #response = HttpResponse(content_type='application/xlsx')
            #response['Content-Disposition'] = 'attachment; filename="{0}"'.format(os.path.basename(tmp.name))
            #response.write(tmp.read())
            #return(response)
            tmp.seek(0)
            request.session['download'] = base64.b64encode(tmp.read())
            download = "true"

        elif request.POST.get('select') == "product":
            wb = load_workbook(request.FILES['product'])
            ws = wb.get_sheet_by_name(name = 'Product_Hierarchy')
            headers = [cell.value for cell in ws.rows.next() ]
            prods=[]
            rowDictList = []

            for row in ws.iter_rows(row_offset=1):
                if row[0].value:
                    rowDict = {}
                    for cellnum in range(len(row) + 1):
                        try:
                            cell = row[cellnum]
                            rowDict[headers[cellnum]] = cell.value
                        except IndexError:
                            pass
                    rowDictList.append(rowDict)


            """All product lines of the same name within a product line request should have the same Igor code, otherwise, abort."""
            productLineGroupIgorDict = {}
            for rowDict in rowDictList:
                productLineGroupIgorDict[rowDict["Product Line Name"]] = productLineGroupIgorDict.get(rowDict["Product Line Name"], []) + [rowDict["Igor Item Class"]]
            for plName, igorCodeList in productLineGroupIgorDict.iteritems():
                if len(set(igorCodeList)) > 1:
                    newClass = "active"
                    messages.warning(request, u"Igor Item Classes differ for identical Product Line Name: {0}. Aborting...".format(plName))
                    return(render(request, "hierarchy/new.html", locals()))

            newProductLines = [] #List of ID's
            subprods = []

            for rowDict in rowDictList:
                try:
                    productlinegroup = ProductLineGroup.objects.get(code__iexact=rowDict.get("Existing Product Line Group Code"))
                except Exception as e:
                    print e
                    print(u"Could not find Product Line Group Code '{0}'".format(rowDict.get("Existing Product Line Group Code")))
                    debug()

                code = get_unused_code()
                igorDescription = rowDict.get("Igor / Sub PL Description", "")
                usage = rowDict.get("Usage")
                productlinename = rowDict.get("Product Line Name")
                if usage:
                    try:
                        usage = Usage.objects.get(name__iexact=usage)
                    except Exception as e:
                        debug()
                        print(u"Usage not found '{0}' at row {1}".format(usage, row[0].row))
                igoritemclass = rowDict.get("Igor Item Class", None)
                if igoritemclass:
                    igoritemclass = IgorItemClass.objects.get(name=igoritemclass)

                """Product line names are not unique in the database, but we don't want to create multiple product lines with the same
                description within a single new product line request. Check if we've created one in this new productline request,
                and if not create a new one. """

                if len(productlinename) > 30:
                    print(u"Product Line Name '{0}' exceed 30 characters: {1} dropping into the debugger. The database will trucate the name if you contine. Abort with ^c".format(productlinename, len(productlinename)))

                try:
                    productLine = ProductLine.objects.get(id__in=newProductLines, name__iexact=productlinename[:30])
                except ProductLine.DoesNotExist:
                    code = get_unused_code()
                    productline = ProductLine(
                        code = code.code,
                        name = productlinename,
                        fproductlinegroup = productlinegroup
                    )
                    productline.save()
                    code.use(newcodes)
                    newProductLines.append(productline.id)
                #Now we have a new productline
                code = get_unused_code()
                subproductline = SubProductLine(
                    fproductline=productline,
                    igor_or_sub_pl=code.code,
                    description=igorDescription,
                    igorclass=igoritemclass,
                    usage=usage
                )
                msg = subproductline.save()
                if msg:
                    messages.warning(request, msg)
                subprods.append(subproductline)
                code.use(newcodes)


            wb = Workbook()
            ws = wb.active
            ws.title = "Product_Hierarchy"
            wsProductCodes = wb.create_sheet()
            wsProductCodes.title = "Product_Codes"
            count=1
            tmp = NamedTemporaryFile(suffix=".xlsx")
            for header in bigheaders:
                ws.cell(row=1, column=count).value= header
                ws.cell(row=1, column=count).font = Font(bold=True)
                count += 1
            count = 2

            for subproduct in subprods:
                subproduct.excel_row(ws, count)
                count += 1

            #tmp.seek(0)
            #response = HttpResponse(content_type='application/xlsx')
            #response['Content-Disposition'] = 'attachment; filename="{0}"'.format(os.path.basename(tmp.name))
            #response.write(tmp.read())
            #return(response)
            tmp.seek(0)
            #Write product codes into a 2nd sheet
            column = 1
            for header in ["Code", "Description", "Type", "Date"]:
                wsProductCodes.cell(row=1, column=column).value = header
                wsProductCodes.cell(row=1, column=column).font = Font(bold=True)
                column += 1

            row = 2
            for newcodeList in newcodes:
                column = 1
                for value in newcodeList:
                    wsProductCodes.cell(row=row, column=column).value = value
                    column += 1
                row += 1
            wb.save(tmp)
            tmp.flush()
            tmp.seek(0)
            request.session['download'] = base64.b64encode(tmp.read())

    elif request.method == "GET":
        download = request.GET.get("download")
        if download in ["sub-product-template.xlsx", "product-template.xlsx"]:
            filepath = os.path.join(settings.BASE_DIR, "hierarchy", "static", "hierarchy", "downloads", download)
            response = HttpResponse(content_type="application/xlsx")
            response['Content-Disposition'] = 'attachment; filename="{0}"'.format(download)
            spreadsheet = open(filepath, 'rb').read()
            response.write(spreadsheet)
            return(response)
        retrieve = "retrieve" in request.GET.keys()
        if retrieve:
            xlsx = base64.b64decode(request.session['download'])
            response = HttpResponse(content_type='application/xlsx')
            response['Content-Disposition'] = 'attachment; filename="{0}"'.format(os.path.basename("productcodes.xlsx"))
            response.write(xlsx)
            return(response)

    newClass = "active"
    return(render(request, "hierarchy/new.html", locals()))

class UploadFileForm(forms.Form):
    file = forms.FileField()

def import_product_hierarchy(request):
    uploadClass = "active"
    form = UploadFileForm()
    if request.method == "POST":
        if request.FILES:
            print("Loading spreadsheet")
            wb = load_workbook(request.FILES['file'])
            ws = wb.get_sheet_by_name(name = 'Product_Hierarchy')
            headers = [cell.value for cell in ws.rows.next() ]
            print("Parsing rows")

            for row in ws.iter_rows(row_offset=1):
                if row[1].value: #if this row isn't empty
                    try:
                        rowDict = {}
                        for cellnum in range(len(row) + 1):
                            try:
                                cell = row[cellnum]
                                rowDict[headers[cellnum]] = cell.value
                            except IndexError:
                                pass
                        segment, created = Segment.objects.get_or_create(name=rowDict["Segment Name"], code=rowDict['Segment Code'])
                        if created:
                            segment.save()
                        division, created = Division.objects.get_or_create(name=rowDict["Division Name"], code=rowDict['Division Code'], fsegment=segment)
                        if created:
                            division.save()
                        businessunit, created = BusinessUnit.objects.get_or_create(name=rowDict["Business Unit Name"], code=rowDict['Business Unit Code'], fdivision = division)
                        if created:
                            businessunit.save()
                        subbusinessunit, created = SubBusinessUnit.objects.get_or_create(name=rowDict["Sub-Business Unit Name"], code=rowDict['Sub-Business Unit Code'], fbusinessunit = businessunit)
                        if created:
                            subbusinessunit.save()
                        name = rowDict["Product Line Group Name"]
                        if len(name) > 30:
                            print(u"Warning: Product Line Group Name '{0}' exceeds 30 characters.".format(name))
                        productlinegroup, created = ProductLineGroup.objects.get_or_create(name=name, code=rowDict['Product Line Group Code'], fsubbusinessunit = subbusinessunit)
                        if created:
                            productlinegroup.save()
                        productline, created = ProductLine.objects.get_or_create(name=rowDict["Product Line Name"], code=rowDict['Product Line Code'], fproductlinegroup = productlinegroup)
                        if created:
                            productline.save()
                        usage, created = Usage.objects.get_or_create(name=rowDict["Usage"])
                        if created:
                            usage.save()

                        igorclassStr = rowDict.get("Igor Item Class")
                        if igorclassStr:
                            try:
                                igorclass = IgorItemClass.objects.get(name=igorclassStr)
                            except Exception as e:
                                print(u"IgorItemClass missing '{0}'".format(igorclassStr))
                                debug()
                        else:
                            igorclass = None

                        usageStr = rowDict.get("Usage")
                        if usageStr:
                            usage = Usage.objects.get(name=usageStr)
                        else:
                            usage = None

                        if rowDict["Igor or Sub PL"]:
                            subproductline, created = SubProductLine.objects.get_or_create(igor_or_sub_pl=rowDict["Igor or Sub PL"], \
                                description=rowDict['Igor / Sub PL Description'], fproductline=productline, usage=usage, igorclass=igorclass)
                            if created:
                                subproductline.save()
                        else:
                            subproductline = None

                        print("."),
                        sys.stdout.flush()

                    except Exception as e:
                        print traceback.format_exc()
                        print("Error on row {0}".format(row[0].row))
                        stop()
            messages.success(request, "Product Hierarchy Loaded")



    return( render(request, 'hierarchy/import_product_hierarchy.html', locals() ))

def download_product_hierarchy(request):
    productlines = ProductLine.objects.all().order_by("id")
    wb = Workbook()
    ws = wb.active
    column = 1
    for header in bigheaders:
        ws.cell(row=1, column=column).value = header
        ws.cell(row=1, column=count).font = Font(bold=True)
        column += 1
    tmp = NamedTemporaryFile(suffix=".xlsx")
    row = 2
    for pl in productlines:
        spls = pl.subproductline_set.all()
        if spls:
            for spl in spls:
                spl.excel_row(ws, row)
                row += 1
        else:
            pl.excel_row(ws, row)
            row += 1
        print("."),
        sys.stdout.flush()
    wb.save(tmp)
    tmp.seek(0)
    response = HttpResponse(content_type='application/xlsx')
    response['Content-Disposition'] = 'attachment; filename="{0}"'.format(os.path.basename(tmp.name))
    response.write(tmp.read())
    return(response)

def aloha(request):

    """Ajax calls to save edited documents"""

    data = {"message":"Unknown Error"}

    if request.is_ajax():
        if request.user.is_superuser:
            #This happens when the editing fairy loads the raw template rather than  the rendered template before editing.
            contentId = request.POST.get('id')
            if contentId:
                templatepath = os.path.join(settings.BASE_DIR, "hierarchy", "templates", "hierarchy", "editable", u"{0}.html".format(contentId))
                try:
                    html = open(templatepath, 'rb').read().decode("utf-8")
                except IOError:
                    html = u""
                return(HttpResponse(html))

            #This happens when the editing fairy is saving a document.
            contentId = request.POST.get(u'contentId')
            content = request.POST.get(u'content')
            templatepath = os.path.join(settings.BASE_DIR, "hierarchy", "templates", "hierarchy", "editable", u"{0}.html".format(contentId))
            if contentId and content:

                """write this document to it's html file on disk."""

                file(templatepath, 'wb').write(content.encode("utf-8"))
                print(u"User {0} updated html file {1}".format(request.user, templatepath))
                data = {}  #Don't pop up any window.
            else:
                data = {}
    else:
        data = {"message":"API"}
    return(JsonResponse(data, safe=False))


class ReplaceCodesForm(forms.Form):
    new_codes_xlsx_file = forms.FileField()


def coderow(row):
    try:
        str(row[2].value)
    except:
        print(u"This entry in col 2 contains unicode: '{0}'".format(row[2].value))
    try:
        str(row[3].value)
    except:
        print(u"This entry in col 2 contains unicode: '{0}'".format(row[3].value))

    code = unicode(row[2].value).upper()
    obj = Code(code=code, used=unicode(row[3].value).upper() == code)
    return(obj)


def replacecodes(request):
    replaceCodesForm = ReplaceCodesForm()
    if request.method=="POST":
        try:
            xlFile=request.FILES['new_codes_xlsx_file']
        except:
            return(render(request, "hierarchy/replacecodes.html", locals()))

        print("Loading workbook...")
        wb = load_workbook(xlFile, data_only=True)

        ws = wb.get_sheet_by_name(name = 'available codes')
        print("Processing codes...")
        coderows = []
        header_row = ws.rows.next()
        for row in ws.iter_rows(row_offset=1):
            coderows.append(coderow(row))
        print("Updating database...")
        deleteResults = Code.objects.all().delete()
        Code.objects.bulk_create(coderows)
        usedCount = Code.objects.filter(used=True).count()
        unusedCount = Code.objects.filter(used=False).count()
        total = Code.objects.all().count()

    return(render(request, "hierarchy/replacecodes.html", locals()))

def update_code(request):
    if request.method == "POST":
        update = request.POST.get("update")
        try:
            if update == "code":
                output = subprocess.Popen(['git', 'pull'], stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
            elif update == "software":
                output = subprocess.Popen(['python', 'update-software.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
        except Exception as e:
            output = ("", str(e))
    return( render(request, 'hierarchy/update_code.html', locals() ))

def reset_database(request):
    dbFile = os.path.join(settings.BASE_DIR, "db.sqlite3")
    removeFiles = [ dbFile]

    for removeFile in removeFiles:
        for filename in glob.glob(removeFile):
            os.remove(filename)
    fromFile = dbFile + ".new"
    shutil.copy(fromFile, dbFile)
    messages.success(request, u"Database reset. Upload new Product Hierarchy and new 3 digit codes.")
    return(redirect(reverse("replacecodes")))